from kivymd.app import MDApp
from kivy.uix.screenmanager import Screen,ScreenManager,NoTransition
import openpyxl
from kivy.lang import Builder
from kivy.storage.jsonstore import JsonStore
from kivymd.uix.label import MDLabel
from kivy.uix.boxlayout import BoxLayout
import json
import os
from kivy.uix.button import Button
from kivymd.uix.card import MDCard
import random
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDRaisedButton,MDFlatButton
from kivy.graphics import Rectangle,Color,Canvas
from kivymd.uix.label import MDLabel
from kivy.uix.image import Image
import threading
help="""
<Screen0>:
	name:"screen_0"
	ScrollView:
    	bar_width:0
		BoxLayout:
			orientation:'vertical'
			height:self.minimum_height+1
			size_hint:(1,None)
			MDIconButton:
				pos_hint:{'center_x':0.5}
				icon:'new-box'
				user_font_size:'300sp'
				on_press:app.dial()
			MDIconButton:
				pos_hint:{'center_x':0.5}
				icon:'folder-open'
				user_font_size:'300sp'
				on_press:app.openfile()
			MDIconButton:
				pos_hint:{'center_x':0.5}
				icon:'ee.png'
				user_font_size:'450sp'
				on_press:
					root.manager.current="screen_6"
			# MDBoxLayout:
			# 	adaptive_height:True
			# 	orientation:'vertical'
			# 	MDLabel:
			# 		text:'Merge online scripts'
			# 		halign:'center'
			# 		bold:True
			# 		font_size:'45sp'
			# 		size_hint_y:self.texture_size[1]
			# 		
			# 	MDIconButton:
			# 		pos_hint:{'center_x':0.5}
			# 		icon:'ee.png'
			# 		user_font_size:'450sp'
			# 		on_press:
			# 			root.manager.current="screen_6"
			
<Screen2>:
	name:"screen_2"
	BoxLayout:
		padding:30
		orientation:"vertical"
		spacing:30
		MDLabel:
			text:"Instructions"
			font_size:35
		TextInput:
			id:m1
			text:""
			font_size:'20sp'
		MDLabel:
			text:"Number of questions"
			font_size:'35sp'
		TextInput:
			id:m2
			text:""
			font_size:'20sp'
		MDLabel:
			text:"Total time"
			font_size:'35sp'
		TextInput:
			id:m3
			text:""
			font_size:'20sp'
		Button:
			text:"Next"
			on_press:app.ssg()
<Screen1>:
	name:"screen_1"
	BoxLayout:
		BoxLayout:
			orientation:"vertical"			
			size_hint_x:0.8
			MDLabel:
				id:Qnum
				text:"1"
				font_style:"H3"
				size_hint_y:None
				height:self.texture_size[1]
			ScrollView:
				bar_width:20
				BoxLayout:
					orientation:"vertical"
					size_hint_y:None
					height:self.minimum_height
					MDLabel:
						halign:"left"
						size_hint:(1,None)
						size:self.texture_size
						text:"Question"
						bold:True
						font_size:'35sp'
					TextInput:
						id:question
						font_size:'20sp'	
						size_hint:(1,None)
						height:self.minimum_height + 100
						background_color:(1,1,1,1)
					MDLabel:
						font_size:'35sp'	
						halign:"left"
						size_hint:(1,None)
						size:self.texture_size
						text:"Correct Answer"
					TextInput:
						id:A1
						font_size:'20sp'
						size_hint:(1,None)
						height:self.minimum_height + 50
						background_color:(1,1,1,1)
					MDLabel:
						font_size:'35sp'
						halign:"left"
						size_hint:(1,None)
						size:self.texture_size
						text:"Wrong Answer"
					TextInput:
						id:A2
						font_size:'20sp'
						size_hint:(1,None)
						height:self.minimum_height + 50
						background_color:(1,1,1,1)
					MDLabel:
						halign:"left"
						font_size:'35sp'
						size_hint:(1,None)
						size:self.texture_size
						text:"Wrong Answer"
					TextInput:
						font_size:'20sp'
						id:A3
						size_hint:(1,None)
						height:self.minimum_height + 50
					MDLabel:
						font_size:'35sp'
						halign:"left"
						size_hint:(1,None)
						size:self.texture_size
						text:"Wrong Answer"
					TextInput:
						font_size:'20sp'
						id:A4
						size_hint:(1,None)
						height:self.minimum_height + 50
						background_color:(1,1,1,1)
			BoxLayout:
				size_hint_y:0.2
				spacing:30
				padding:30
				Button:
                    text:'Previous'
                    size_hint_x:.5 
                    bold:True 
                    on_press:app.previous_quest()          
				Button:
                    size_hint_x:.5 
                    bold:True					
            		text:"Next"
					on_press:app.next_quest()
		BoxLayout:
			size_hint_x:0.2
			id:bx
			orientation:"vertical"
			spacing:30
			padding:15
			MDIconButton:
				pos_hint:{'center_x':0.5}
				icon:'content-save'
				user_font_size:'130sp'
				on_press:app.save_quest()
			MDIconButton:
				pos_hint:{'center_x':0.5}
				icon:'eye'
				user_font_size:'130sp'
				on_press:app.view_quest()
			
			Widget:
			
<Screen4>:
	name:"screen_4"
	FileChooserIconView:
		id:filechooser
    	on_selection:app.selected(filechooser.selection)
    

<Screen4b>:
	name:"screen_4b"
	FileChooserIconView:
    	id:filechooserr
    	on_selection:app.selectedb(filechooserr.selection)
<Screen4c>:
	name:"screen_4c"
	FileChooserIconView:
		id:filechooserrr
		on_selection:app.selectedc(filechooserrr.selection)
<Screen5>:
	name:"screen_5"
	BoxLayout:
		orientation:'vertical'
		MDIconButton:
			icon: "arrow-left"
			#pos_hint: {"center_x": .04, "center_y":.95}
			valign:'left'
			user_font_size:'50sp'
			on_press:app.re()
		ScrollView:
			BoxLayout:
				padding:20
				spacing:10
				size_hint_y:None
				height:self.minimum_height+1000
				id:scview
				orientation:"vertical"
	
<Screen6>:
	name:"screen_6"
	BoxLayout:
        orientation:'vertical'
        size_hint:(.7,.7)
        pos_hint:{'center_x':0.5,'center_y':0.5}
        spacing:100
        BoxLayout:
            spacing:20
            MDTextField:
                id:fil
                background_active:self.background_normal
            MDRaisedButton:
            	id:filr
                theme_text_color:'Custom'
                bold:True
                font_size:'30sp' 
                md_bg_color:(0,0,0,1)
                md_bg_color_disabled:(0,0,0,1)
                text:'Select guide'
                on_press:app.choosefile()
        BoxLayout:
            spacing:20
            MDTextField:
                id:fill
                background_active:self.background_normal
            MDRaisedButton:
            	id:filt
                theme_text_color:'Custom'
                bold:True
                font_size:'30sp' 
                md_bg_color:(0,0,0,1)
                md_bg_color_disabled:(0,0,0,1)
                text:'Select student scripts'
                on_press:app.choosefilee()
        MDRaisedButton:
        	id:fily
            text:'Generate results'
            bold:True
            font_size:'60sp' 
            pos_hint:{'center_x':0.5}
            md_bg_color:(0,0,0,1)
            md_bg_color_disabled:(0,0,0,1)
            on_press:root.generate_results()
<Screen7>:
	name:"screen_7"
	ScrollView:
    	#bar_width:'20dp'
    	do_scroll_x:False
    	do_scroll_y:True
		MDGridLayout:
			cols:2
			spacing:'20dp'
			height:self.minimum_height+7
			size_hint:(1,None)
			id:gr1
			MDLabel:
				text:'CANDIDATE'
				font_style:'H4'
				bold:True
			MDLabel:
				text:'SCORE'
				font_style:'H4'
				bold:True
				# size_hint:(None,1)
				# width:self.minimum_size[0]
<Content>:
	size_hint_y:None
	spacing:20
	height:150
	MDTextField:
		id:ss1
		text:''
		background_active:self.background_normal
<Contentt>:
	size_hint_y:None
	spacing:20
	height:150
	MDTextField:
		id:ss2
		text:''
		background_active:self.background_normal
			"""
class Screen0(Screen):
	pass
class Screen1(Screen):
	def on_enter(self):
		self.ids.Qnum.text=MainApp.selected.rr
		with open(Screen1.hh,"r") as file:
			existing_que=json.load(file)
			Questions=existing_que["questions"]
class Screen2(Screen):
	pass
# class Screen3(Screen):
# 	pass
class Screen4(Screen):
	pass
class Screen4b(Screen):
	pass
class Screen4c(Screen):
	pass
class Screen5(Screen):
	pass
class Screen6(Screen):
	def generate_results(self):
		sm.get_screen("screen_6").ids.fil.disabled = True
		sm.get_screen("screen_6").ids.fill.disabled = True
		sm.get_screen("screen_6").ids.filr.disabled = True
		sm.get_screen("screen_6").ids.filt.disabled = True
		sm.get_screen("screen_6").ids.fily.disabled = True
		with self.canvas:
			Color(.5, 0.5, 0.5, 0.4)
			self.rect = Rectangle(pos=self.center, size=(self.width, self.height))
			self.rect.pos = self.pos
			self.rect.size = self.size
			self.bind(pos=self.update_rect,size=self.update_rect)
		ll=MDLabel(text='Processing...',font_style='H2',
													 halign='center',pos_hint={'center_y':.8}, bold=True)
		sm.get_screen("screen_6").ids['ll']=ll
		sm.get_screen("screen_6").add_widget(ll)
		# mdc=MDCard(size_hint=(.2,.2))
		im=Image(source='YouTube_loading_symbol_3_(transparent) (1).gif',
			  size_hint=(.2, .2), pos_hint={'center_x': .5, 'center_y': .2})
		sm.get_screen("screen_6").ids['iol'] =im
		# mdc.add_widget(Image(source='image_processing20210902-19849-1kf32tm.gif',pos_hint={'center_x':.6,'center_y':.7}))
		sm.get_screen("screen_6").add_widget(im)
		res = sm.get_screen("screen_6").ids.fil.text
		ress = sm.get_screen("screen_6").ids.fill.text
		with open(res, "r") as file:
			restt = json.load(file)
		with open(ress, "r") as file:
			rest = json.load(file)
		for e in rest:
			print(e)
			score = 0
			for f in e['questt']:
				z = e['questt'][f]['q']
				i = 0
				try:
					while restt["questions"][i]["Q"] != z:
						i = i + 1
					if e['questt'][f]['ans'] == restt["questions"][i]["Ans"][0]:
						score = score + 1
					else:
						pass
				except:
					pass
			print(e['candidte'])
			print(score)
			scores[e['candidte']]=  score
		sm.get_screen("screen_6").remove_widget(sm.get_screen("screen_6").ids['iol'])
		sm.get_screen("screen_6").ids.ll.text='Done'
		he=MDRaisedButton(text='Save as .xls', theme_text_color='Custom', md_bg_color=(0, 0, 0, 1),
						  pos_hint={'center_x':0.5,'center_y':0.6},font_size='50sp',on_press=MainApp.make_xls)
		sm.get_screen("screen_6").ids['xls'] = he
		sm.get_screen("screen_6").add_widget(he)
		hee = MDRaisedButton(text='Save as .db', theme_text_color='Custom', md_bg_color=(0, 0, 0, 1),
							pos_hint={'center_x': 0.5, 'center_y': 0.4}, font_size='50sp')
		sm.get_screen("screen_6").ids['xls'] = hee
		sm.get_screen("screen_6").add_widget(hee)
		print(scores)
	def update_rect(self,*args):
		self.rect.pos=self.pos
		self.rect.size=self.size
class Screen7(Screen):
	pass
class Content(BoxLayout):
	pass
class Contentt(BoxLayout):
	pass
class MainApp(MDApp):
	global sm,Questions,pressedtimes,scores
	sm=ScreenManager()
	Questions=[]
	pressedtimes=[]
	scores={}
	def build(self):
		self.theme_cls.theme_style="Dark"
		Builder.load_string(help)
		sm.add_widget(Screen0(name="screen_0"))
		sm.add_widget(Screen1(name="screen_1"))
		sm.add_widget(Screen2(name="screen_2"))
		#sm.add_widget(Screen3(name="screen_3"))
		sm.add_widget(Screen4(name="screen_4"))
		sm.add_widget(Screen4b(name="screen_4b"))
		sm.add_widget(Screen4c(name="screen_4c"))
		sm.add_widget(Screen5(name="screen_5"))
		sm.add_widget(Screen6(name="screen_6"))
		sm.add_widget(Screen7(name="screen_7"))
		return sm
	# def cc(self):
	# 	th1.start()
	def make_xls(self):
		MainApp.dialogg = MDDialog(title='Enter filename', type='custom', content_cls=Contentt(),
							   buttons=[MDRaisedButton(text='save', on_press=MainApp.xll)])
		MainApp.dialogg.open()
	def xll(self):
		ww = MainApp.dialogg.content_cls.ids.ss2.text
		par = os.getcwd()
		new_d = 'Scores'
		new_path = os.path.join(par, new_d)
		try:
			os.mkdir(new_path)
			xl_files = openpyxl.Workbook(f'{new_path}/{ww}.xlsx')
			xl_files.save(f'{new_path}/{ww}.xlsx')
			xl_files = openpyxl.load_workbook(f'{new_path}/{ww}.xlsx')
			sheet=xl_files['Sheet']
			sheet.cell(1,1).value='Candidate'
			sheet.cell(1, 2).value = 'Score'
			row=2
			col=1
			for gra in scores:
				sheet.cell(row,col).value=gra
				col=col+1
				sheet.cell(row, col).value = scores[gra]
				row=row+1
				col=1
			xl_files.save(f'{new_path}/{ww}.xlsx')
		except:
			xl_files = openpyxl.Workbook(f'{new_path}/{ww}.xlsx')
			xl_files.save(f'{new_path}/{ww}.xlsx')
			xl_files = openpyxl.load_workbook(f'{new_path}/{ww}.xlsx')
			sheet = xl_files['Sheet']
			sheet.cell(1, 1).value = 'Candidate'
			sheet.cell(1, 2).value = 'Score'
			row = 2
			col = 1
			for gra in scores:
				sheet.cell(row, col).value = gra
				col = col + 1
				sheet.cell(row, col).value = scores[gra]
				row = row + 1
				col = 1
			xl_files.save(f'{new_path}/{ww}.xlsx')

		MainApp.dialogg.dismiss()
		sm.current = "screen_6"
	def dial(self):
		self.dialog= MDDialog(title='Enter filename',type='custom',content_cls=Content(),
							  buttons=[MDRaisedButton(text='save',on_press=self.ppp)])
		self.dialog.open()
	def choosefile(self):
		sm.current = "screen_4b"
	def choosefilee(self):
		sm.current = "screen_4c"
	def ssg(self):
		with open(self.file_to_work,"r") as file:
			ert=json.load(file)
			ert["instructions"]=self.root.get_screen("screen_2").ids.m1.text
			ert["number of questions"]=int(self.root.get_screen("screen_2").ids.m2.text)
			ert["time"]=int(self.root.get_screen("screen_2").ids.m3.text)
		with open(self.file_to_work,"w") as file:
			json.dump(ert,file)
		MainApp.selected.rr='1'
		sm.current="screen_1"
	def ppp(self,obj):
		ww=self.dialog.content_cls.ids.ss1.text
		fp = open(f'{ww}.json', 'w')
		fp.write('{"instructions": "", "number of questions": "", "time": "", "questions": []}')
		fp.close()
		self.file_to_work=f'{ww}.json'
		self.ooo=self.file_to_work[0:-5]
		self.dialog.dismiss()
		sm.current="screen_2"
	def selected(self,filename):
		print(filename[0])
		if ".json" in filename[0]:
			try:
				self.file_to_work=filename[0]
				self.ooo=self.file_to_work[0:-5]
				Screen1.hh=filename[0]
				with open(self.file_to_work, "r") as file:
					get_quantity = json.load(file)
					MainApp.selected.rr=str(len(get_quantity["questions"])+1)
				for element in get_quantity["questions"]:
					Questions.append(element)
				sm.current="screen_1"
			except:
				pass
		else:
			pass
	def selectedb(self,filename):
		if ".json" in filename[0]:
			self.ref=filename[0]
			#self.ooo=self.ref[0:-5]
			self.root.get_screen("screen_6").ids.fil.text=self.ref
			sm.current="screen_6"
		else:
			pass
	def selectedc(self,filename):
		if ".json" in filename[0]:
			self.reff=filename[0]
			#self.ooo=self.ref[0:-5]
			self.root.get_screen("screen_6").ids.fill.text=self.reff
			sm.current="screen_6"
		else:
			pass
	def previous_quest(self):
		if  int(self.root.get_screen("screen_1").ids.Qnum.text) - 1>= 1:
			self.root.get_screen("screen_1").ids.Qnum.text = str(int(self.root.get_screen("screen_1").ids.Qnum.text) - 1)
			self.root.get_screen("screen_1").ids.question.text=Questions[int(self.root.get_screen("screen_1").ids.Qnum.text) - 1]['Q']
			self.root.get_screen("screen_1").ids.A1.text=Questions[int(self.root.get_screen("screen_1").ids.Qnum.text) - 1]['Ans'][0]
			self.root.get_screen("screen_1").ids.A2.text =Questions[int(self.root.get_screen("screen_1").ids.Qnum.text) - 1]['Ans'][1]
			self.root.get_screen("screen_1").ids.A3.text =Questions[int(self.root.get_screen("screen_1").ids.Qnum.text) - 1]['Ans'][3]
			self.root.get_screen("screen_1").ids.A4.text =Questions[int(self.root.get_screen("screen_1").ids.Qnum.text) - 1]['Ans'][3]
		else:
			pass
	def openfile(self):
		sm.current="screen_4"
	def next_quest(self):
		if int(self.root.get_screen("screen_1").ids.Qnum.text) >= len(Questions):
			Questions.append({"Q":self.root.get_screen("screen_1").ids.question.text,"Ans":(self.root.get_screen("screen_1").ids.A1.text,self.root.get_screen("screen_1").ids.A2.text,self.root.get_screen("screen_1").ids.A3.text,self.root.get_screen("screen_1").ids.A4.text)})
			self.root.get_screen("screen_1").ids.Qnum.text=str(int(self.root.get_screen("screen_1").ids.Qnum.text)+1)
			self.root.get_screen("screen_1").ids.question.text=""
			self.root.get_screen("screen_1").ids.A1.text=""
			self.root.get_screen("screen_1").ids.A2.text=""
			self.root.get_screen("screen_1").ids.A3.text=""
			self.root.get_screen("screen_1").ids.A4.text=""
		else:
			self.root.get_screen("screen_1").ids.Qnum.text = str(int(self.root.get_screen("screen_1").ids.Qnum.text) + 1)
			self.root.get_screen("screen_1").ids.question.text = Questions[int(self.root.get_screen("screen_1").ids.Qnum.text)-1]['Q']
			self.root.get_screen("screen_1").ids.A1.text = Questions[int(self.root.get_screen("screen_1").ids.Qnum.text)-1]['Ans'][0]
			self.root.get_screen("screen_1").ids.A2.text = Questions[int(self.root.get_screen("screen_1").ids.Qnum.text)-1]['Ans'][1]
			self.root.get_screen("screen_1").ids.A3.text = Questions[int(self.root.get_screen("screen_1").ids.Qnum.text)-1]['Ans'][2]
			self.root.get_screen("screen_1").ids.A4.text = Questions[int(self.root.get_screen("screen_1").ids.Qnum.text)-1]['Ans'][3]
	def save_quest(self):
		with open(self.file_to_work,"r") as file:
			existing_que=json.load(file)
			existing_que["questions"]=Questions
		with open(self.file_to_work,"w") as file:
			json.dump(existing_que, file)
	def view_quest(self):
		sm.current="screen_5"
		i=0
		z=0
		while i< len(Questions):
		#for e in Questions:
			btn=MDLabel(text=f"Q{i+1}",font_style='H4',bold=True,halign='center')
			self.root.get_screen("screen_5").ids[f'{z}'] = btn
			z=z+1
			self.root.get_screen("screen_5").ids.scview.add_widget(btn)
			btn=MDLabel(text=Questions[i]["Q"],bold=True,font_style='H4')
			self.root.get_screen("screen_5").ids[f'{z}'] = btn
			z = z + 1
			self.root.get_screen("screen_5").ids.scview.add_widget(btn)
			btn=MDLabel(text=Questions[i]["Ans"][0],font_style='H6')
			self.root.get_screen("screen_5").ids[f'{z}'] = btn
			z = z + 1
			self.root.get_screen("screen_5").ids.scview.add_widget(btn)
			btn=MDLabel(text=Questions[i]["Ans"][1],font_style='H6')
			self.root.get_screen("screen_5").ids[f'{z}'] = btn
			z = z + 1
			self.root.get_screen("screen_5").ids.scview.add_widget(btn)
			btn=MDLabel(text=Questions[i]["Ans"][2],font_style='H6')
			self.root.get_screen("screen_5").ids[f'{z}'] = btn
			z = z + 1
			self.root.get_screen("screen_5").ids.scview.add_widget(btn)
			btn=MDLabel(text=Questions[i]["Ans"][3],font_style='H6')
			self.root.get_screen("screen_5").ids[f'{z}'] = btn
			z = z + 1
			self.root.get_screen("screen_5").ids.scview.add_widget(btn)
			btn = MDLabel(text='',size_hint_y=None,height=20,font_style='H6')
			self.root.get_screen("screen_5").ids[f'{z}'] = btn
			z = z + 1
			self.root.get_screen("screen_5").ids.scview.add_widget(btn)
			i=i+1
		self.ple=z

	def re(self):
		for i in range(0,self.ple):
			self.root.get_screen("screen_5").ids.scview.remove_widget(self.root.get_screen("screen_5").ids[f'{i}'])
		sm.current = "screen_1"

#
# th1=threading.Thread(target=MainApp.cc, args=(1,))
		
MainApp().run()